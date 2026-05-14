#!/usr/bin/env python3
"""
writers/excel.py — xlsx 导出器（无截图，报表用途）。

依赖: openpyxl

用法:
  python3 -m writers.excel --input output/zxgk_batch_20260514_zhixing.json
  python3 -m writers.excel --input output/zxgk_batch_20260514_zhixing.json --output report.xlsx
"""

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("错误: 需要安装 openpyxl — pip install openpyxl", file=sys.stderr)
    sys.exit(1)


HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)


def write_sheet(wb, sheet_name, data):
    ws = wb.create_sheet(title=sheet_name)
    headers = ["被执行人", "案号", "立案日期", "viewId", "截图路径"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for company_data in data.get("companies", []):
        for rec in company_data.get("records", []):
            ws.cell(row=row, column=1, value=rec.get("name", ""))
            ws.cell(row=row, column=2, value=rec.get("caseNo", ""))
            ws.cell(row=row, column=3, value=rec.get("date", ""))
            ws.cell(row=row, column=4, value=rec.get("viewId", ""))
            ws.cell(row=row, column=5, value=rec.get("screenshot", ""))
            row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 40
    return row - 2


def write_batch(json_files, output_path):
    """将一个或多个 batch JSON 写入 xlsx，每个 subsite 一个 sheet"""
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认 sheet

    total = 0
    for json_path in json_files:
        with open(json_path) as f:
            data = json.load(f)
        subsite = data.get("subsite", Path(json_path).stem)
        sheet_name = subsite[:31]  # Excel sheet name max 31 chars
        n = write_sheet(wb, sheet_name, data)
        print(f"Excel: {n} 条写入 sheet '{sheet_name}'")
        total += n

    wb.save(output_path)
    print(f"Excel: 共 {total} 条 → {output_path}")
    return total


def main():
    parser = argparse.ArgumentParser(description="Excel 导出器")
    parser.add_argument("--input", required=True, nargs="+", help="batch JSON 文件路径（可多个）")
    parser.add_argument("--output", default=None, help="输出 xlsx 路径")
    args = parser.parse_args()

    if args.output is None:
        # 从第一个 input 文件名推断日期
        stem = Path(args.input[0]).stem
        args.output = f"{stem}.xlsx"

    for p in args.input:
        if not Path(p).exists():
            print(f"错误: 文件不存在 — {p}", file=sys.stderr)
            sys.exit(1)

    write_batch(args.input, args.output)


if __name__ == "__main__":
    main()
